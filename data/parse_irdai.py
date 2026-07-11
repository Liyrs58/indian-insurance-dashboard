#!/usr/bin/env python3
import json
from datetime import datetime
import os
import re

import openpyxl

RESEARCH_AS_OF = "2026-07-07"
SOURCE_LINKS = [
    {
        "name": "General Insurance Council Flash Figures",
        "url": "https://www.gicouncil.in/statistics/industry-statistics/flash-figures/",
        "latest_observed": "Flash Figures June 2026, published Tuesday 07 July, 2026",
    },
    {
        "name": "IRDAI Reports & Statistics - Monthly business figures",
        "url": "https://irdai.gov.in/non-life",
        "latest_observed": "IRDAI monthly business figures archive",
    },
    {
        "name": "Life Insurance Council - New Business Performance",
        "url": "https://www.lifeinscouncil.org/industry%20information/nbp.aspx",
        "latest_observed": "Life council archive page available; local raw files currently through May 2026",
    },
]

MONTH_NAME_TO_NUM = {
    'january': '01', 'jan': '01', 'जनवरी': '01',
    'february': '02', 'feb': '02', 'फ़रवरी': '02', 'फरवरी': '02',
    'march': '03', 'mar': '03', 'मार्च': '03',
    'april': '04', 'apr': '04', 'अप्रैल': '04',
    'may': '05', 'मई': '05',
    'june': '06', 'jun': '06', 'जून': '06',
    'july': '07', 'jul': '07', 'जुलाई': '07',
    'august': '08', 'aug': '08', 'अगस्त': '08',
    'september': '09', 'sep': '09', 'सितंबर': '09', 'सितम्बर': '09',
    'october': '10', 'oct': '10', 'अक्टूबर': '10', 'अक्तूबर': '10',
    'november': '11', 'nov': '11', 'नवंबर': '11', 'नवम्बर': '11',
    'december': '12', 'dec': '12', 'दिसंबर': '12', 'दिसम्बर': '12',
}

# Hindi to English mapping for non-life insurers
NON_LIFE_NAMES = {
    'एको जनरल इंश्योरेंस लिमिटेड': 'Acko General Insurance',
    'बजाज जनरल इंश्योरेंस लिमिटेड': 'Bajaj Allianz General Insurance',
    'बजाज एलियांज जनरल इंश्योरेंस लिमिटेड': 'Bajaj Allianz General Insurance',
    'चोलमांडलम एमएस जनरल इंश्योरेंस कंपनी लिमिटेड': 'Cholamandalam MS General Insurance',
    'जनरली सेंट्रल इंश्योरेंस कंपनी लिमिटेड': 'General Central Insurance',
    'गो डिजिट जनरल इंश्योरेंस लिमिटेड': 'Go Digit General Insurance',
    'एचडीएफसी एर्गो जनरल इंश्योरेंस कंपनी लिमिटेड': 'HDFC ERGO General Insurance',
    'आईसीआईसीआई लोम्बार्ड जनरल इंश्योरेंस कंपनी लिमिटेड': 'ICICI Lombard General Insurance',
    'इफको-टोकियो जनरल इंश्योरेंस कंपनी लिमिटेड': 'IFFCO-Tokio General Insurance',
    'क्षेमा जनरल इंश्योरेंस कंपनी लिमिटेड': 'Kshema General Insurance',
    'लिबर्टी जनरल इंश्योरेंस लिमिटेड': 'Liberty General Insurance',
    'मैग्मा जनरल इंश्योरेंस लिमिटेड': 'Magma General Insurance',
    'नेशनल इंश्योरेंस कंपनी लिमिटेड': 'National Insurance Company',
    'नवी जनरल इंश्योरेंस लिमिटेड': 'Navi General Insurance',
    'रहेजा क्यूबीई जनरल इंश्योरेंस कंपनी लिमिटेड': 'Raheja QBE General Insurance',
    'इंडसइंड जनरल इंश्योरेंस कंपनी लिमिटेड': 'IndusInd General Insurance',
    'रॉयल सुंदरम जनरल इंश्योरेंस कंपनी लिमिटेड': 'Royal Sundaram General Insurance',
    'एसबीआई जनरल इंश्योरेंस कंपनी लिमिटेड': 'SBI General Insurance',
    'श्रीराम जनरल इंश्योरेंस कंपनी लिमिटेड': 'Shriram General Insurance',
    'टाटा एआईजी जनरल इंश्योरेंस कंपनी लिमिटेड': 'Tata AIG General Insurance',
    'न्यू इंडिया एश्योरेंस कंपनी लिमिटेड': 'New India Assurance',
    'ओरिएंटल इंश्योरेंस कंपनी लिमिटेड': 'Oriental Insurance',
    'यूनाइटेड इंडिया इंश्योरेंस कंपनी लिमिटेड': 'United India Insurance',
    'यूनिवर्सल सोपो जनरल इंश्योरेंस कंपनी लिमिटेड': 'Universal Sompo General Insurance',
    'जूनो जनरल इंश्योरेंस कंपनी लिमिटेड': 'Zuno General Insurance',
    'ज्यूरिक कोटक जनरल इंश्योरेंस कंपनी लिमिटेड': 'Zurich Kotak General Insurance',
    'आदित्य बिड़ला हेल्थ इंश्योरेंस कंपनी लिमिटेड': 'Aditya Birla Health Insurance',
    'केयर हेल्थ इंश्योरेंस लिमिटेड': 'Care Health Insurance',
    'मनीपालसिग्ना हेल्थ इंश्योरेंस कंपनी लिमिटेड': 'ManipalCigna Health Insurance',
    'निवा बूपा हेल्थ इंश्योरेंस कंपनी लिमिटेड': 'Niva Bupa Health Insurance',
    'रिलायंस हेल्थ इंश्योरेंस लिमिटेड*': 'Reliance Health Insurance',
    'स्टार हेल्थ & एलाइड इंश्योरेंस कंपनी लिमिटेड': 'Star Health & Allied Insurance',
    'नारायणा हेल्थ इंश्योरेंस लिमिटेड': 'Narayana Health Insurance',
    'गैलेक्सी हेल्थ इंश्योरेंस कंपनी लिमिटेड': 'Galaxy Health Insurance',
    'एग्रीकल्चर इंश्योरेंस कम्पनी ऑफ इंडिया लिमिटेड': 'Agriculture Insurance Company of India',
    'ईसीजीसी लिमिटेड': 'ECGC Limited',
    # Additional mappings for names found in the data
    'रिलायंस जनरल इंश्योरेंस कंपनी लिमिटेड': 'Reliance General Insurance',
    'यूनिवर्सल सोम्पो जनरल इंश्योरेंस कंपनी लिमिटेड': 'Universal Sompo General Insurance',
    'एमजीआईजी जनरल इंश्योरेंस कंपनी लिमिटेड': 'Magma General Insurance',
    'एनआईसी लिमिटेड': 'National Insurance Company',
    'यूआईजी लिमिटेड': 'United India Insurance',
    'ओआईसी लिमिटेड': 'Oriental Insurance',
    'न्यू इंडिया एश्योरेंस कंपनी लिमिटेड': 'New India Assurance',
    'निवा बूपा हेल्थ इंश्योरेंस कंपनी लिमिटेड': 'Niva Bupa Health Insurance',
    'एचडीएफसी एर्गो जनरल इंश्योरेंस कंपनी लिमिटेड': 'HDFC ERGO General Insurance',
    'स्टार हेल्थ एंड एलाइड इंश्योरेंस कंपनी लिमिटेड': 'Star Health & Allied Insurance',
    # More variations found in actual data
    'बजाज एलियांज जनरल इंश्योरेंस कंपनी लिमिटेड': 'Bajaj Allianz General Insurance',
    'रिलायंस जनरल इंश्योरेंस': 'Reliance General Insurance',
    'बजाज एलियांज जनरल इंश्योरेंस': 'Bajaj Allianz General Insurance',
    'नारायना हेल्थ इंश्योरेंस लिमिटेड': 'Narayana Health Insurance',
}

# Hindi to English mapping for life insurers
LIFE_NAMES = {
    'एको लाइफ इंश्योरेंस लिमिटेड': 'Acko Life Insurance',
    'आदित्य बिड़ला सन लाइफ इंश्योरेंस कंपनी लिमिटेड': 'Aditya Birla Sun Life Insurance',
    'एजियास फेडरल लाइफ इंश्योरेंस कंपनी लिमिटेड': 'Ageas Federal Life Insurance',
    'अवीवा लाइफ इंश्योरेंस कंपनी इंडिया लिमिटेड': 'Aviva Life Insurance',
    'बजाज आलियांज लाइफ इंश्योरेंस कंपनी लिमिटेड': 'Bajaj Allianz Life Insurance',
    'बंधन लाइफ इंश्योरेंस लिमिटेड': 'Bandhan Life Insurance',
    '\nभारती एक्सा लाइफ इंश्योरेंस कंपनी लिमिटेड': 'Bharti AXA Life Insurance',
    'भारती एक्सा लाइफ इंश्योरेंस कंपनी लिमिटेड': 'Bharti AXA Life Insurance',
    'एचडीएफसी लाइफ इंश्योरेंस कंपनी लिमिटेड': 'HDFC Life Insurance',
    'आईसीआईसीआई प्रुडेंशियल लाइफ इंश्योरेंस कंपनी लिमिटेड': 'ICICI Prudential Life Insurance',
    'इंडियाफर्स्ट लाइफ इंश्योरेंस कंपनी लिमिटेड': 'IndiaFirst Life Insurance',
    'कोटक महिंद्रा लाइफ इंश्योरेंस कंपनी लिमिटेड': 'Kotak Mahindra Life Insurance',
    'कोटक लाइफ इंश्योरेंस कंपनी लिमिटेड': 'Kotak Life Insurance',
    'एलआईसी (भारतीय जीवन बीमा निगम)': 'LIC',
    'भारतीय जीवन बीमा निगम': 'LIC',
    'एक्सिस मैक्स लाइफ इंश्योरेंस लिमिटेड': 'Max Life Insurance',
    'मैक्स लाइफ इंश्योरेंस कंपनी लिमिटेड': 'Max Life Insurance',
    'एमपीएल लाइफ इंश्योरेंस कंपनी लिमिटेड': 'MPL Life Insurance',
    'पीएनबी मेटलाइफ इंडिया इंश्योरेंस कंपनी लिमिटेड': 'PNB MetLife India Insurance',
    'पीएनबी मेटलाइफ इंश्योरेंस कंपनी लिमिटेड': 'PNB MetLife Insurance',
    'प्राइम एशिया लाइफ इंश्योरेंस कंपनी लिमिटेड': 'Prime Asia Life Insurance',
    'प्रामेरिका लाइफ इंश्योरेंस लिमिटेड': 'Pramerica Life Insurance',
    'इंडसइंड निप्पॉन लाइफ इंश्योरेंस कंपनी लिमिटेड': 'IndusInd Nippon Life Insurance',
    'सहारा लाइफ इंश्योरेंस कंपनी लिमिटेड': 'Sahara Life Insurance',
    'एसबीआई लाइफ इंश्योरेंस कंपनी लिमिटेड': 'SBI Life Insurance',
    'श्रीराम लाइफ इंश्योरेंस कंपनी लिमिटेड': 'Shriram Life Insurance',
    'स्टार यूनियन दाई-इची लाइफ इंश्योरेंस कंपनी लिमिटेड': 'Star Union Dai-ichi Life Insurance',
    'टाटा एआईए लाइफ इंश्योरेंस कंपनी लिमिटेड': 'Tata AIA Life Insurance',
    'फ्यूचर जनरली इंडिया लाइफ इंश्योरेंस कंपनी लिमिटेड': 'Future Generali India Life Insurance',
    'रिलायंस निप्पॉन लाइफ इंश्योरेंस कंपनी लिमिटेड': 'Reliance Nippon Life Insurance',
    # Additional mappings for names found in actual data
    'केनरा एचएसबीसी लाइफ इंश्योरेंस कंपनी लिमिटेड': 'Canara HSBC Life Insurance',
    'क्रेडिटएक्सेस लाइफ इंश्योरेंस लिमिटेड': 'CreditAccess Life Insurance',
    'एडलवाइस लाइफ इंश्योरेंस कंपनी लिमिटेड': 'Edelweiss Life Insurance',
    'जनरली सेंट्रल लाइफ इंश्योरेंस कंपनी लिमिटेड': 'Generali Central Life Insurance',
    'गो डिजिट लाइफ इंश्योरेंस लिमिटेड': 'Go Digit Life Insurance',
    'आईसीआईसीआई प्रूडेंशियल लाइफ इंश्योरेंस कंपनी लिमिटेड': 'ICICI Prudential Life Insurance',
}

def parse_month_from_filename(filename):
    """Extract month from filename like NonLife_GDP_March2026.xlsx"""
    filename_lower = filename.lower()

    for month_name, month_num in MONTH_NAME_TO_NUM.items():
        if month_name in filename_lower:
            year_match = re.search(r'(\d{4})', filename)
            if year_match:
                return f"{year_match.group(1)}-{month_num}"
    
    return None

def parse_month_from_text(text):
    if not text:
        return None
    normalized = re.sub(r"\s+", " ", str(text)).strip()
    lower = normalized.lower()
    years = [int(y) for y in re.findall(r'(20\d{2})', normalized)]
    years = [y for y in years if 2000 <= y <= 2035]
    if not years:
        return None
    year = max(years)
    for month_name, month_num in MONTH_NAME_TO_NUM.items():
        if month_name.isascii():
            if re.search(rf'\b{re.escape(month_name)}\b', lower):
                return f"{year}-{month_num}"
        elif month_name in lower:
            return f"{year}-{month_num}"
    return None

def parse_month_from_worksheet(ws):
    parts = []
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        for value in row:
            if value:
                parts.append(str(value))
    return parse_month_from_text(" ".join(parts))

def resolve_report_month(filename, ws=None):
    filename_month = parse_month_from_filename(filename)
    header_month = parse_month_from_worksheet(ws) if ws else None
    month = header_month or filename_month
    if header_month and filename_month and header_month != filename_month:
        print(f"  WARNING: {filename} filename month {filename_month} differs from workbook header {header_month}; using header")
    return month, filename_month, header_month

def safe_float(value, default=0):
    """Safely convert a value to float"""
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if not text or text in {"-", "—", "–"}:
            return default
        try:
            negative = text.startswith("(") and text.endswith(")")
            cleaned = text.replace(",", "").replace(" ", "").replace("%", "")
            cleaned = cleaned.strip("()")
            if cleaned:
                num = float(cleaned)
                return -num if negative else num
        except:
            pass
    return default

def pct_float(value, default=0):
    return safe_float(value, default)

def fiscal_year_for_month(month):
    year = int(month[:4])
    month_num = int(month[5:7])
    end_year = year + 1 if month_num >= 4 else year
    return f"FY{end_year - 1}-{str(end_year)[-2:]}"

def month_label(month):
    return datetime.strptime(month, "%Y-%m").strftime("%b %Y")

def normalize_english_name(name):
    """Normalize insurer names so cross-month lookups and ranking labels are stable."""
    if not name:
        return name
    n = re.sub(r"\s+", " ", str(name)).strip()
    n = n.replace("&amp;", "&")
    replacements = {
        "Acko General Insurance Ltd": "Acko General Insurance",
        "Acko General Insurance Limited": "Acko General Insurance",
        "Acko Life Insurance": "Acko Life Insurance Limited",
        "Bajaj General Insurance Limited": "Bajaj Allianz General Insurance",
        "Bajaj Allianz General Insurance Limited": "Bajaj Allianz General Insurance",
        "Cholamandalam MS General Insurance Co Ltd": "Cholamandalam MS General Insurance",
        "Generali Central Insurance Company Limited": "General Central Insurance",
        "Go Digit General Insurance Ltd": "Go Digit General Insurance",
        "Go Digit General Insurance Limited": "Go Digit General Insurance",
        "HDFC Ergo General Insurance Co Ltd": "HDFC ERGO General Insurance",
        "HDFC ERGO General Insurance Company Limited": "HDFC ERGO General Insurance",
        "ICICI Lombard General Insurance Co Ltd": "ICICI Lombard General Insurance",
        "ICICI Lombard General Insurance Company Limited": "ICICI Lombard General Insurance",
        "IFFCO-Tokio General Insurance Co Ltd": "IFFCO-Tokio General Insurance",
        "IndusInd General Insurance Company Limited": "IndusInd General Insurance",
        "Kiwi General Insurance Ltd": "Kiwi General Insurance",
        "Kshema General insurance": "Kshema General Insurance",
        "Liberty General Insurance Co. Ltd": "Liberty General Insurance",
        "Magma General Insurance Limited": "Magma General Insurance",
        "National Insurance Co Ltd": "National Insurance Company",
        "Navi General Insurance Co. Ltd": "Navi General Insurance",
        "Raheja QBE General Insurance Co Ltd": "Raheja QBE General Insurance",
        "Royal Sundaram General Insurance Co Ltd": "Royal Sundaram General Insurance",
        "SBI General Insurance Co Ltd": "SBI General Insurance",
        "Shriram General Insurance Co Ltd": "Shriram General Insurance",
        "Tata AIG General Insurance Co Ltd": "Tata AIG General Insurance",
        "The New India Assurance Co Ltd": "New India Assurance",
        "The Oriental Insurance Co Ltd": "Oriental Insurance",
        "United India Insurance Co Ltd": "United India Insurance",
        "Universal Sompo General Insurance Co Ltd": "Universal Sompo General Insurance",
        "Zuno General Insurance Co Ltd": "Zuno General Insurance",
        "Zurich Kotak Mahindra General Insurance Co Ltd": "Zurich Kotak General Insurance",
        "Niva bupa health insurance company limited": "Niva Bupa Health Insurance",
        "Aditya Birla Health Insurance Co Ltd": "Aditya Birla Health Insurance",
        "Care Health Insurance Ltd": "Care Health Insurance",
        "Galaxy Health Insurance Company Ltd": "Galaxy Health Insurance",
        "ManipalCigna Health Insurance Co Ltd": "ManipalCigna Health Insurance",
        "Narayana Health Insurance Ltd": "Narayana Health Insurance",
        "Star Health & Allied Insurance Co Ltd": "Star Health & Allied Insurance",
        "Agriculture Insurance Co Of India Ltd": "Agriculture Insurance Company of India",
        "ECGC Ltd": "ECGC Limited",
        "Life Insurance Corporation of India": "Life Insurance Corporation of India",
        "LIC": "Life Insurance Corporation of India",
    }
    return replacements.get(n, n)

def sort_insurers(insurers):
    return sorted(insurers, key=lambda x: x.get("premium_cr", 0), reverse=True)

def aggregate_growth(current_total, prior_total):
    if prior_total:
        return ((current_total - prior_total) / prior_total) * 100
    return 0

def build_month_record(month, filename, insurers, total_premium_cr, total_growth_pct, period_basis="cumulative_ytd", **extra):
    record = {
        "month": month,
        "month_label": month_label(month),
        "fiscal_year": fiscal_year_for_month(month),
        "period_type": period_basis,
        "source_file": filename,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "insurers": sort_insurers(insurers),
        "total_premium_cr": round(total_premium_cr, 2),
        "total_growth_pct": round(total_growth_pct, 2),
    }
    record.update({k: v for k, v in extra.items() if v is not None})
    return record

def month_record_score(record):
    score = 0
    if record.get('filename_month') == record.get('month'):
        score += 2
    if record.get('header_month') == record.get('month'):
        score += 1
    if record.get('month_source') == 'workbook_header':
        score += 1
    return score

def dedupe_month_records(records, segment_name, duplicate_resolutions=None):
    by_month = {}
    for record in records:
        month = record['month']
        current = by_month.get(month)
        if not current:
            by_month[month] = record
            continue

        keep, drop = current, record
        if month_record_score(record) > month_record_score(current):
            keep, drop = record, current
        reason = (
            "higher month-source confidence"
            if month_record_score(keep) != month_record_score(drop)
            else "first record retained on equal confidence"
        )
        note = (
            f"Duplicate {segment_name} month {month}: kept {keep['source_file']} "
            f"and dropped {drop['source_file']}"
        )
        print(f"  WARNING: {note}")
        notes = keep.setdefault('extraction_notes', [])
        notes.append(note)
        if duplicate_resolutions is not None:
            duplicate_resolutions.append({
                "segment": segment_name,
                "month": month,
                "kept_source_file": keep['source_file'],
                "dropped_source_file": drop['source_file'],
                "reason": reason,
            })
        by_month[month] = keep

    return sorted(by_month.values(), key=lambda x: x['month'])

def build_source_hygiene(raw_files_processed, loaded_records, retained_records, duplicate_resolutions):
    mismatches = []
    for record in loaded_records:
        filename_month = record.get('filename_month')
        header_month = record.get('header_month')
        if filename_month and header_month and filename_month != header_month:
            mismatches.append({
                "source_file": record.get('source_file'),
                "filename_month": filename_month,
                "header_month": header_month,
                "selected_month": record.get('month'),
            })

    return {
        "raw_files_processed": raw_files_processed,
        "records_loaded": len(loaded_records),
        "records_retained": len(retained_records),
        "records_dropped": max(0, len(loaded_records) - len(retained_records)),
        "filename_header_mismatches": mismatches,
        "duplicate_resolutions": duplicate_resolutions,
    }

def parse_non_life_excel(filepath, filename):
    """Parse non-life Excel file and extract insurer data"""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[wb.sheetnames[0]]
    
    month, filename_month, header_month = resolve_report_month(filename, ws)
    if not month:
        print(f"Warning: Could not parse month from {filename}")
        return None
    
    insurers = []
    total_premium_cr = 0
    total_growth_pct = 0
    
    for i, row in enumerate(ws.iter_rows(max_row=100, values_only=True)):
        # Data rows start from row 12 (index 11)
        if i >= 11:
            # Check if it's a data row (has serial number in column A)
            serial = row[0]
            if serial and isinstance(serial, (int, float)) and serial > 0:
                hindi_name = str(row[1]).strip() if row[1] else ''
                # Remove any leading/trailing whitespace and newlines
                hindi_name = re.sub(r'\s+', ' ', hindi_name).strip()
                
                # Get premium values with safe conversion
                cumulative_2025_26 = safe_float(row[4])
                cumulative_2024_25 = safe_float(row[5])
                
                # Calculate growth
                if cumulative_2024_25 > 0:
                    growth_pct = ((cumulative_2025_26 - cumulative_2024_25) / cumulative_2024_25) * 100
                else:
                    growth_pct = 0
                
                # Map to English name
                english_name = translate_hindi_name(hindi_name, NON_LIFE_NAMES)
                
                # Clean up the name
                english_name = english_name.replace('लिमिटेड', 'Limited').replace('कंपनी', 'Company')
                english_name = normalize_english_name(english_name)
                
                insurers.append({
                    'name': english_name,
                    'premium_cr': round(cumulative_2025_26, 2),
                    'prior_premium_cr': round(cumulative_2024_25, 2),
                    'market_share_pct': 0,  # Will be calculated later
                    'yoy_growth_pct': round(growth_pct, 2)
                })
                
                total_premium_cr += cumulative_2025_26
        
        # Stop at total row
        if row[1] and 'कुल योग' in str(row[1]):
            break
    
    # Calculate market shares
    if total_premium_cr > 0:
        for insurer in insurers:
            insurer['market_share_pct'] = round((insurer['premium_cr'] / total_premium_cr) * 100, 2)
    
    prior_total = sum(i.get('prior_premium_cr', 0) for i in insurers)
    total_growth_pct = aggregate_growth(total_premium_cr, prior_total)
    
    wb.close()

    # Validate share sum
    share_sum = sum(i['market_share_pct'] for i in insurers)
    if share_sum < 99.0 or share_sum > 101.0:
        print(f"  WARNING: Non-Life share sum = {share_sum:.2f}% (expected ~100%)")

    # Check for negative premiums
    negative_prems = [i for i in insurers if i['premium_cr'] < 0]
    if negative_prems:
        for n in negative_prems:
            print(f"  NOTE: Signed negative source premium for {n['name']}: {n['premium_cr']} Cr")

    # Flag extreme growth values
    extreme = [i for i in insurers if abs(i['yoy_growth_pct']) > 300]
    if extreme:
        for e in extreme:
            print(f"  NOTE: Extreme growth for {e['name']}: {e['yoy_growth_pct']:.1f}% (low base)")

    return build_month_record(
        month, filename, insurers, total_premium_cr, total_growth_pct,
        filename_month=filename_month,
        header_month=header_month,
        month_source="workbook_header" if header_month else "filename",
    )

def translate_hindi_name(hindi_name, name_mapping):
    """Try to translate Hindi name to English using partial matching"""
    # First try exact match
    if hindi_name in name_mapping:
        return name_mapping[hindi_name]
    
    # Try partial matching
    for hindi, english in name_mapping.items():
        # Check if the key contains the input or vice versa
        if hindi in hindi_name or hindi_name in hindi:
            return english
    
    # Try to match by common insurance company name patterns
    common_patterns = {
        'एलआईसी': 'LIC',
        'लिमिटेड': 'Limited',
        'कंपनी': 'Company',
        'इंश्योरेंस': 'Insurance',
        'लाइफ': 'Life',
    }
    
    # Return the original name if no translation found
    return hindi_name

def parse_life_excel(filepath, filename):
    """Parse life Excel file and extract insurer data"""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[wb.sheetnames[0]]
    
    month, filename_month, header_month = resolve_report_month(filename, ws)
    if not month:
        print(f"Warning: Could not parse month from {filename}")
        return None
    
    insurers = []
    total_premium_cr = 0
    total_growth_pct = 0
    
    # Life insurance data rows are every 7 rows (main row + 6 sub-rows for premium types)
    # Main rows are at: 4, 11, 18, 25, 32, 39, 46, 53, 60, 67, 74, 81, 88, 95, 102, 109, 116, 123, 130, 137, 144, 151, 158, 165, 179
    main_rows = [4, 11, 18, 25, 32, 39, 46, 53, 60, 67, 74, 81, 88, 95, 102, 109, 116, 123, 130, 137, 144, 151, 158, 165, 179]
    
    for i, row in enumerate(ws.iter_rows(max_row=200, values_only=True), 1):
        # Only process main data rows
        if i not in main_rows:
            continue
        
        # Data rows have serial number in column A
        serial = row[0]
        if serial and isinstance(serial, (int, float)) and serial > 0:
            hindi_name = str(row[1]).strip() if row[1] else ''
            hindi_name = re.sub(r'\s+', ' ', hindi_name).strip()
            
            # For life insurance:
            # Column C (index 2): March 2025 monthly premium
            # Column D (index 3): March 2026 monthly premium
            # Column E (index 4): Growth rate %
            # Column F (index 5): March 2025 cumulative premium
            # Column G (index 6): March 2026 cumulative premium
            # Column H (index 7): Growth rate %
            # Column I (index 8): Market share %
            
            premium_cumulative_2026 = safe_float(row[6])  # Column G
            premium_cumulative_2025 = safe_float(row[5])  # Column F
            market_share = safe_float(row[8])  # Column I
            
            # Calculate YoY growth
            if premium_cumulative_2025 > 0:
                growth_pct = ((premium_cumulative_2026 - premium_cumulative_2025) / premium_cumulative_2025) * 100
            else:
                growth_pct = 0
            
            # Map to English name
            english_name = translate_hindi_name(hindi_name, LIFE_NAMES)
            
            # Clean up the name
            english_name = english_name.replace('लिमिटेड', 'Limited').replace('कंपनी', 'Company')
            english_name = normalize_english_name(english_name)
            
            # Market share in Excel is already in percentage format
            # (e.g., 2.74 means 2.74%, not 0.0274)
            market_share_pct = round(market_share, 2)
            
            # Only add if premium is > 0 (skip rows with 0 premium)
            if premium_cumulative_2026 > 0:
                insurers.append({
                    'name': english_name,
                    'premium_cr': round(premium_cumulative_2026, 2),
                    'prior_premium_cr': round(premium_cumulative_2025, 2),
                    'market_share_pct': market_share_pct,
                    'yoy_growth_pct': round(growth_pct, 2)
                })
                
                total_premium_cr += premium_cumulative_2026
    
    prior_total = sum(i.get('prior_premium_cr', 0) for i in insurers)
    total_growth_pct = aggregate_growth(total_premium_cr, prior_total)
    
    wb.close()

    # Validate share sum
    share_sum = sum(i['market_share_pct'] for i in insurers)
    if share_sum < 99.0 or share_sum > 101.0:
        print(f"  WARNING: Life share sum = {share_sum:.2f}% (expected ~100%)")

    # Check for negative premiums
    negative_prems = [i for i in insurers if i['premium_cr'] < 0]
    if negative_prems:
        for n in negative_prems:
            print(f"  NOTE: Signed negative source premium for {n['name']}: {n['premium_cr']} Cr")

    # Flag extreme growth values
    extreme = [i for i in insurers if abs(i['yoy_growth_pct']) > 300]
    if extreme:
        for e in extreme:
            print(f"  NOTE: Extreme growth for {e['name']}: {e['yoy_growth_pct']:.1f}% (low base)")

    return build_month_record(
        month, filename, insurers, total_premium_cr, total_growth_pct,
        filename_month=filename_month,
        header_month=header_month,
        month_source="workbook_header" if header_month else "filename",
    )

def parse_non_life_pdf(filepath, filename):
    """Parse General Insurance Council non-life flash PDF exports."""
    try:
        import pdfplumber
    except ImportError:
        print(f"Warning: pdfplumber unavailable, skipping {filename}")
        return None

    month = parse_month_from_filename(filename)
    if not month:
        print(f"Warning: Could not parse month from {filename}")
        return None

    insurers = []
    total_premium_cr = 0
    prior_total = 0

    skip_names = {
        "General Insurers",
        "Stand Alone Health Insurers",
        "Specialised Insurers",
        "General Insurers Sub Total",
        "Stand Alone Health Insurers sub Total",
        "Specialised Insurers Sub Total",
        "Grand Total",
        "Grand Total excl. specialised cos",
        "Insurers",
        "",
    }

    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                for row in table:
                    if not row or not row[0]:
                        continue
                    raw_name = re.sub(r"\s+", " ", str(row[0])).strip()
                    if raw_name in skip_names or raw_name.startswith("Flash Report") or raw_name.startswith("Gross Direct"):
                        continue
                    if len(row) < 8:
                        continue

                    cumulative_current = safe_float(row[4])
                    cumulative_previous = safe_float(row[5])
                    if cumulative_current == 0 and cumulative_previous == 0:
                        continue

                    growth_pct = pct_float(row[6])
                    if cumulative_previous and not row[6]:
                        growth_pct = aggregate_growth(cumulative_current, cumulative_previous)
                    market_share = pct_float(row[7], None)
                    if market_share is None:
                        market_share = 0

                    insurers.append({
                        "name": normalize_english_name(raw_name),
                        "premium_cr": round(cumulative_current, 2),
                        "prior_premium_cr": round(cumulative_previous, 2),
                        "market_share_pct": round(market_share, 2),
                        "yoy_growth_pct": round(growth_pct, 2),
                    })
                    total_premium_cr += cumulative_current
                    prior_total += cumulative_previous

    if not insurers:
        print(f"Warning: No insurer rows parsed from {filename}")
        return None

    total_growth_pct = aggregate_growth(total_premium_cr, prior_total)
    filename_month = parse_month_from_filename(filename)
    return build_month_record(
        month, filename, insurers, total_premium_cr, total_growth_pct,
        filename_month=filename_month,
        month_source="filename",
    )

def build_validation(life_data, non_life_data):
    issues = []

    def add(severity, code, message, **extra):
        item = {"severity": severity, "code": code, "message": message}
        item.update(extra)
        issues.append(item)

    def is_fiscal_year_reset(prev, curr):
        return (
            prev.get('fiscal_year') != curr.get('fiscal_year')
            and prev.get('period_type') == 'cumulative_ytd'
            and curr.get('period_type') == 'cumulative_ytd'
        )

    for segment_name, segment_data in [('life', life_data), ('non_life', non_life_data)]:
        for month_data in segment_data:
            share_sum = sum(i.get('market_share_pct') or 0 for i in month_data['insurers'])
            if share_sum < 99.0 or share_sum > 101.0:
                add(
                    "warning",
                    "share_sum",
                    f"{segment_name} {month_data['month']} market shares sum to {share_sum:.2f}%",
                    segment=segment_name,
                    month=month_data['month'],
                    value=round(share_sum, 2),
                    source_file=month_data.get('source_file'),
                )

            for insurer in month_data['insurers']:
                if insurer.get('premium_cr', 0) < 0 or insurer.get('market_share_pct', 0) < 0:
                    add(
                        "info",
                        "signed_source_adjustment",
                        f"{insurer['name']} has a signed negative source premium/share in {segment_name} {month_data['month']}; retained as reported",
                        segment=segment_name,
                        month=month_data['month'],
                        insurer=insurer['name'],
                        premium_cr=insurer.get('premium_cr'),
                        market_share_pct=insurer.get('market_share_pct'),
                        source_file=month_data.get('source_file'),
                    )
                if abs(insurer.get('yoy_growth_pct') or 0) > 300:
                    add(
                        "info",
                        "extreme_growth",
                        f"{insurer['name']} has extreme YoY growth in {segment_name} {month_data['month']}",
                        segment=segment_name,
                        month=month_data['month'],
                        insurer=insurer['name'],
                        yoy_growth_pct=insurer.get('yoy_growth_pct'),
                    )

        for i in range(1, len(segment_data)):
            prev = segment_data[i - 1]
            curr = segment_data[i]
            if curr['total_premium_cr'] == prev['total_premium_cr']:
                add(
                    "error",
                    "duplicate_total",
                    f"{segment_name} months {prev['month']} and {curr['month']} have identical total premium",
                    segment=segment_name,
                    months=[prev['month'], curr['month']],
                    value=curr['total_premium_cr'],
                )
            if prev['total_premium_cr'] and curr['total_premium_cr'] / prev['total_premium_cr'] < 0.5:
                if is_fiscal_year_reset(prev, curr):
                    add(
                        "info",
                        "fiscal_year_reset",
                        f"{segment_name} cumulative YTD premium resets between fiscal years",
                        segment=segment_name,
                        from_month=prev['month'],
                        to_month=curr['month'],
                        from_value=prev['total_premium_cr'],
                        to_value=curr['total_premium_cr'],
                        from_fiscal_year=prev.get('fiscal_year'),
                        to_fiscal_year=curr.get('fiscal_year'),
                    )
                else:
                    add(
                        "warning",
                        "large_period_drop",
                        f"{segment_name} premium drops more than 50%; check for source-period mismatch",
                        segment=segment_name,
                        from_month=prev['month'],
                        to_month=curr['month'],
                        from_value=prev['total_premium_cr'],
                        to_value=curr['total_premium_cr'],
                    )

    life_months = {d['month'] for d in life_data}
    non_life_months = {d['month'] for d in non_life_data}
    for month in sorted(non_life_months - life_months):
        add("info", "missing_companion_month", f"Non-life has {month}, but life data is unavailable", month=month, missing_segment="life")
    for month in sorted(life_months - non_life_months):
        add("info", "missing_companion_month", f"Life has {month}, but non-life data is unavailable", month=month, missing_segment="non_life")

    status = "ok"
    if any(i["severity"] == "error" for i in issues):
        status = "error"
    elif any(i["severity"] == "warning" for i in issues):
        status = "warning"

    return {
        "status": status,
        "issue_count": len(issues),
        "issues": issues,
    }

def write_analysis_summary(output_data, output_path):
    life = output_data['life']['monthly_data'][-1]
    non_life = output_data['non_life']['monthly_data'][-1]
    shared_month = output_data['_meta'].get('latest_shared_month')
    comparable = output_data['summary']

    def line_for(insurer, idx):
        return f"{idx}. {insurer['name']}: ₹{insurer['premium_cr']/1000:.1f}K Cr ({insurer['market_share_pct']:.1f}% share, {insurer['yoy_growth_pct']:+.1f}% YoY)"

    lines = [
        "# IRDAI Market Analysis Summary",
        "",
        "## Market Overview",
        f"- Comparable Month: {shared_month}",
        f"- Total Comparable Premium: ₹{comparable['total_market_premium_cr']/1000:.1f}K Cr",
        f"- Life Premium: ₹{comparable['life_premium_cr']/1000:.1f}K Cr",
        f"- Non-Life Premium: ₹{comparable['non_life_premium_cr']/1000:.1f}K Cr",
        f"- Insurance Penetration: {comparable['insurance_penetration_pct']}%",
        f"- Insurance Density: ${comparable['insurance_density_usd']}",
        "",
        f"## Top Life Insurers (Latest Life Month: {life['month']})",
    ]
    lines.extend(line_for(i, idx) for idx, i in enumerate(life['insurers'][:5], 1))
    lines.extend(["", f"## Top Non-Life Insurers (Latest Non-Life Month: {non_life['month']})"])
    lines.extend(line_for(i, idx) for idx, i in enumerate(non_life['insurers'][:5], 1))
    lines.extend([
        "",
        "## Data Caveats",
        "- Figures are provisional and unaudited where source flash reports say so.",
        "- Cumulative YTD figures reset at fiscal-year boundaries; avoid reading fiscal resets as market collapses.",
        "- Source-hygiene metadata records raw files processed, duplicate resolutions, and filename/header month mismatches.",
        "- Validation issues and source caveats are stored in data/irdai-processed.json under _meta.validation.",
    ])
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("\n".join(lines) + "\n")

def main():
    data_dir = '/Users/rudra/Documents/New OpenCode Project/insurance-dashboard/data/irdai-excel'
    output_file = '/Users/rudra/Documents/New OpenCode Project/insurance-dashboard/data/irdai-processed.json'
    existing_file = '/Users/rudra/Documents/New OpenCode Project/insurance-dashboard/data/irdai-data.json'
    
    # Read existing data if available
    existing_data = {}
    if os.path.exists(existing_file):
        with open(existing_file, 'r', encoding='utf-8') as f:
            existing_data = json.load(f)
    
    non_life_data = []
    life_data = []
    raw_files_processed = 0
    
    # Process local raw files. XLSX files come from the original IRDAI-style archive;
    # the July 7 non-life flash release is currently published by GIC as PDF.
    for filename in os.listdir(data_dir):
        if filename.endswith('.xlsx') or filename.endswith('.pdf'):
            filepath = os.path.join(data_dir, filename)
            raw_files_processed += 1
            print(f"Processing: {filename}")
            
            if filename.endswith('.pdf') and 'NonLife' in filename:
                data = parse_non_life_pdf(filepath, filename)
                if data:
                    non_life_data.append(data)
                    print(f"  Found {len(data['insurers'])} insurers, Total: {data['total_premium_cr']} Cr")
            elif 'NonLife' in filename:
                data = parse_non_life_excel(filepath, filename)
                if data:
                    non_life_data.append(data)
                    print(f"  Found {len(data['insurers'])} insurers, Total: {data['total_premium_cr']} Cr")
            elif 'Life' in filename:
                data = parse_life_excel(filepath, filename)
                if data:
                    life_data.append(data)
                    print(f"  Found {len(data['insurers'])} insurers, Total: {data['total_premium_cr']} Cr")
    
    loaded_records = non_life_data + life_data
    duplicate_resolutions = []

    # Resolve duplicate records after header-based month detection. This catches
    # mislabeled local files before they can create fake periods in the UI.
    non_life_data = dedupe_month_records(non_life_data, 'non_life', duplicate_resolutions)
    life_data = dedupe_month_records(life_data, 'life', duplicate_resolutions)
    retained_records = non_life_data + life_data

    # Get all available months
    all_months = sorted(set([d['month'] for d in non_life_data] + [d['month'] for d in life_data]))
    shared_months = sorted(set([d['month'] for d in non_life_data]) & set([d['month'] for d in life_data]))

    life_by_month = {d['month']: d for d in life_data}
    non_life_by_month = {d['month']: d for d in non_life_data}

    # Calculate summary from the latest shared month only. Do not combine
    # different segment dates in the headline market figure.
    latest_non_life = non_life_data[-1] if non_life_data else None
    latest_life = life_data[-1] if life_data else None
    latest_shared_month = shared_months[-1] if shared_months else None
    comparable_life = life_by_month.get(latest_shared_month) if latest_shared_month else latest_life
    comparable_non_life = non_life_by_month.get(latest_shared_month) if latest_shared_month else latest_non_life

    life_premium = comparable_life['total_premium_cr'] if comparable_life else 0
    non_life_premium = comparable_non_life['total_premium_cr'] if comparable_non_life else 0
    total_market_premium = life_premium + non_life_premium
    validation = build_validation(life_data, non_life_data)
    source_hygiene = build_source_hygiene(
        raw_files_processed,
        loaded_records,
        retained_records,
        duplicate_resolutions,
    )
    
    # Create output JSON
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    output_data = {
        "_meta": {
            "source": "IRDAI Flash Figures",
            "generated_at": now_str,
            "last_updated": datetime.now().strftime("%Y-%m-%d"),
            "research_as_of": RESEARCH_AS_OF,
            "months_available": all_months,
            "shared_months": shared_months,
            "latest_shared_month": latest_shared_month,
            "latest_life_month": latest_life['month'] if latest_life else None,
            "latest_non_life_month": latest_non_life['month'] if latest_non_life else None,
            "source_links": SOURCE_LINKS,
            "validation": validation,
            "source_hygiene": source_hygiene,
            "notes": "Data extracted from local IRDAI/GIC flash source files. Headline market totals use the latest shared life/non-life month only."
        },
        "non_life": {
            "monthly_data": non_life_data
        },
        "life": {
            "monthly_data": life_data
        },
        "summary": {
            "total_market_premium_cr": round(total_market_premium, 2),
            "life_premium_cr": round(life_premium, 2),
            "non_life_premium_cr": round(non_life_premium, 2),
            "insurance_penetration_pct": existing_data.get('market_overview', {}).get('insurance_penetration', {}).get('value', 3.7),
            "insurance_density_usd": existing_data.get('market_overview', {}).get('insurance_density', {}).get('value', 92),
            "global_penetration_avg_pct": 7.0
        }
    }
    
    # Merge with existing data if available
    if existing_data:
        # Keep existing data that's not in the new file
        for key in existing_data:
            if key not in output_data:
                output_data[key] = existing_data[key]
    
    # Write output JSON
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, indent=2, ensure_ascii=False)

    write_analysis_summary(output_data, os.path.join(os.path.dirname(output_file), 'analysis_summary.md'))
    
    print(f"\nData saved to: {output_file}")
    print(f"Total non-life insurers: {len(latest_non_life['insurers']) if latest_non_life else 0}")
    print(f"Total life insurers: {len(latest_life['insurers']) if latest_life else 0}")
    print(f"Comparable market premium ({latest_shared_month}): {total_market_premium:.2f} Cr")
    print(f"Validation status: {validation['status']} ({validation['issue_count']} issues)")

if __name__ == '__main__':
    main()
